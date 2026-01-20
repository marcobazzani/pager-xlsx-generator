#!/usr/bin/env python3
"""
PagerDuty-style on-call shift scheduler that generates an xlsx file with rotating shifts.
Loads schedule configuration from external YAML files.

Each layer in the config represents a single time window (e.g., 08:00-10:30).
Multiple layers compose the complete daily coverage.
The output includes all layers with exact dates calculated from start to end date.

Usage:
    python oncall_scheduler.py --config schedule_2_5h.yaml
    python oncall_scheduler.py --config schedule_2_5h.yaml --start-date 2026-02-01 --end-date 2026-05-01
    python oncall_scheduler.py --config schedule_2_5h.yaml --generate-ics
"""

import argparse
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import yaml
import os
import re
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import Rectangle
import matplotlib.dates as mdates


def parse_date_argument(date_str, reference_date=None):
    """
    Parse a date argument that can be either absolute (YYYY-MM-DD) or relative (+Nd, +Nw, +Nm, +Ny).
    
    Args:
        date_str: Date string in format YYYY-MM-DD or relative format (+2d, +3w, +2m, +1y)
        reference_date: Reference date for relative calculations (defaults to today)
    
    Returns:
        datetime object
    
    Examples:
        "2026-01-20" -> Jan 20, 2026
        "+2d" or "+2" -> 2 days from reference date
        "+3w" -> 3 weeks from reference date
        "+2m" -> 2 months from reference date
        "+1y" -> 1 year from reference date
        "today" -> today's date
    """
    if reference_date is None:
        reference_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    date_str = date_str.strip().lower()
    
    # Handle "today" keyword
    if date_str == "today":
        return reference_date
    
    # Handle relative dates: +Nd, +Nw, +Nm, +Ny (or just +N for days)
    relative_pattern = r'^\+(\d+)([dwmy])?$'
    match = re.match(relative_pattern, date_str)
    
    if match:
        amount = int(match.group(1))
        unit = match.group(2) or 'd'  # Default to days if no unit specified
        
        if unit == 'd':
            return reference_date + timedelta(days=amount)
        elif unit == 'w':
            return reference_date + timedelta(weeks=amount)
        elif unit == 'm':
            return reference_date + relativedelta(months=amount)
        elif unit == 'y':
            return reference_date + relativedelta(years=amount)
    
    # Try parsing as absolute date (YYYY-MM-DD)
    try:
        return datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        raise ValueError(
            f"Invalid date format '{date_str}'. "
            f"Use YYYY-MM-DD (e.g., 2026-01-20) or relative format "
            f"(e.g., +2d, +3w, +2m, +1y, or today)"
        )



def load_schedule_config(config_file):
    """
    Load schedule configuration from YAML file.
    
    Args:
        config_file: Path to YAML configuration file
    
    Returns:
        Dictionary with schedule configuration
    """
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"Configuration file not found: {config_file}")
    
    with open(config_file, 'r') as f:
        config = yaml.safe_load(f)
    
    if 'schedule' not in config:
        raise ValueError(f"Invalid configuration file: missing 'schedule' key")
    
    return config['schedule']


def calculate_date_range(schedule_config, start_date_override=None, end_date_override=None):
    """
    Calculate the date range for the schedule.
    
    Args:
        schedule_config: Schedule configuration dictionary
        start_date_override: Optional start date override (datetime)
        end_date_override: Optional end date override (datetime)
    
    Returns:
        Tuple of (start_date, end_date) as datetime objects
    """
    if start_date_override:
        start_date = start_date_override
    else:
        # Use config default or today
        config_start = schedule_config.get('start_date')
        if config_start:
            start_date = datetime.strptime(config_start, '%Y-%m-%d')
        else:
            start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    if end_date_override:
        end_date = end_date_override
    else:
        # Use config duration or default 3 months
        duration_months = schedule_config.get('duration_months', 3)
        end_date = start_date + relativedelta(months=duration_months)
    
    return start_date, end_date


def generate_dates_for_layer(layer_config, start_date, end_date):
    """
    Generate all dates where a layer should have shifts.
    Now supports time_windows per day configuration.
    
    Args:
        layer_config: Layer configuration dictionary
        start_date: Start date (datetime)
        end_date: End date (datetime)
    
    Returns:
        List of tuples (datetime, day_name) for dates where this layer is active
    """
    day_map = {
        'monday': 0, 'tuesday': 1, 'wednesday': 2, 
        'thursday': 3, 'friday': 4, 'saturday': 5, 'sunday': 6
    }
    
    # New structure: time_windows with per-day configuration
    if 'time_windows' in layer_config:
        time_windows = layer_config.get('time_windows', {})
        active_weekdays = {day_map[day.lower()]: day.lower() 
                          for day in time_windows.keys() if day.lower() in day_map}
    else:
        # Fallback to old structure for backward compatibility
        active_days = layer_config.get('days', [])
        active_weekdays = {day_map[day.lower()]: day.lower() 
                          for day in active_days if day.lower() in day_map}
    
    dates = []
    current_date = start_date
    
    while current_date < end_date:
        weekday = current_date.weekday()
        if weekday in active_weekdays:
            dates.append((current_date, active_weekdays[weekday]))
        current_date += timedelta(days=1)
    
    return dates


def generate_oncall_calendar(config_file, output_file, start_date_override=None, end_date_override=None):
    """
    Generate the on-call calendar xlsx file in PagerDuty style.
    All layers are included in a single output file with exact dates.
    
    Args:
        config_file: Path to YAML configuration file
        output_file: Output xlsx filename
        start_date_override: Optional start date override (datetime)
        end_date_override: Optional end date override (datetime)
    """
    # Load configuration
    schedule_config = load_schedule_config(config_file)
    schedule_name = schedule_config.get('name', 'On-Call Schedule')
    schedule_description = schedule_config.get('description', '')
    
    # Calculate date range
    start_date, end_date = calculate_date_range(schedule_config, start_date_override, end_date_override)
    
    layers_config = schedule_config.get('layers', {})
    
    if not layers_config:
        raise ValueError("No layers defined in configuration file")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "On-Call Schedule"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    layer_header_fill = PatternFill(start_color="36C5F0", end_color="36C5F0", fill_type="solid")
    layer_header_font = Font(bold=True, color="000000", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Color palette for different people
    colors = [
        "E8F5E9", "E3F2FD", "FFF3E0", "FCE4EC", "F3E5F5",
        "E0F2F1", "FFF9C4", "FFE0B2", "F8BBD0", "D1C4E9",
        "C8E6C9", "BBDEFB", "FFE0B2", "F8BBD0", "E1BEE7"
    ]
    
    # Write header
    row = 1
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = schedule_name
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.alignment = center_align
    
    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = schedule_description
    cell.alignment = center_align
    cell.font = Font(italic=True)
    
    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    cell.alignment = center_align
    cell.font = Font(bold=True)
    
    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    cell.alignment = center_align
    cell.font = Font(size=9, italic=True)
    
    row += 2
    
    # Process each layer
    layer_shifts = []  # List of (date, layer_name, time_window, person, layer_id)
    
    for layer_idx, (layer_id, layer_config) in enumerate(layers_config.items()):
        layer_name = layer_config.get('name', layer_id)
        rotation_team = layer_config.get('rotation_team', [])
        is_dummy_layer = layer_config.get('dummy', False)  # Check if entire layer is dummy
        
        if not rotation_team:
            continue
        
        # Check if using new time_windows structure or old time_window
        time_windows = layer_config.get('time_windows', {})
        old_time_window = layer_config.get('time_window', {})
        
        # Generate dates for this layer
        dates = generate_dates_for_layer(layer_config, start_date, end_date)
        
        # Assign people in rotation
        for date_idx, (shift_date, day_name) in enumerate(dates):
            person = rotation_team[date_idx % len(rotation_team)]
            
            # Get time window for this specific day
            is_dummy_day = False
            if time_windows and day_name in time_windows:
                day_window = time_windows[day_name]
                start_time = day_window.get('start', 'N/A')
                end_time = day_window.get('end', 'N/A')
                is_dummy_day = day_window.get('dummy', False)  # Check if specific day is dummy
            else:
                # Fallback to old structure
                start_time = old_time_window.get('start', 'N/A')
                end_time = old_time_window.get('end', 'N/A')
            
            # Skip adding dummy shifts to the output (either layer-level or day-level dummy)
            if not is_dummy_layer and not is_dummy_day:
                layer_shifts.append((
                    shift_date,
                    layer_name,
                    f"{start_time} - {end_time}",
                    person,
                    layer_idx
                ))
    
    # Sort shifts by date and then by start time
    layer_shifts.sort(key=lambda x: (x[0], x[2]))
    
    # Write column headers (removed Layer column, added Start Time, End Time, Hours, On-Call Status)
    headers = ['Date', 'Day', 'Start Time', 'End Time', 'Hours', 'On-Call Person', 'On-Call Status']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    row += 1
    start_data_row = row
    
    # Track person colors
    person_colors = {}
    color_idx = 0
    
    # Write data rows
    current_date = None
    for shift_date, layer_name, time_window, person, layer_idx in layer_shifts:
        # Add empty row between different dates
        if current_date and current_date != shift_date:
            row += 1  # Empty row separator
        
        current_date = shift_date
        
        # Assign color to person if not yet assigned
        if person not in person_colors:
            person_colors[person] = colors[color_idx % len(colors)]
            color_idx += 1
        
        color = person_colors[person]
        
        # Parse time window to get start and end times
        time_parts = time_window.split(' - ')
        start_time = time_parts[0] if len(time_parts) > 0 else 'N/A'
        end_time = time_parts[1] if len(time_parts) > 1 else 'N/A'
        
        # Write row data (removed Layer column)
        ws.cell(row=int(row), column=1).value = shift_date  # Date as datetime for formula
        ws.cell(row=int(row), column=2).value = shift_date.strftime('%A')
        ws.cell(row=int(row), column=3).value = start_time
        ws.cell(row=int(row), column=4).value = end_time
        
        # Add Hours formula in column 5 (D-C converted to hours)
        row_num = int(row)
        hours_formula = f'=(D{row_num}-C{row_num})*24'
        ws.cell(row=row_num, column=5).value = hours_formula
        
        ws.cell(row=int(row), column=6).value = person
        
        # Add On-Call Status formula in column 7
        # Formula checks if NOW() is between Date+StartTime and Date+EndTime
        oncall_formula = f'=IF(AND(NOW()>=A{row_num}+C{row_num},NOW()<=A{row_num}+D{row_num}),"On-Call","")'
        ws.cell(row=row_num, column=7).value = oncall_formula
        
        # Apply styling
        for col in range(1, 8):
            cell = ws.cell(row=int(row), column=col)
            cell.border = border
            cell.alignment = center_align
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Format date column as date
        ws.cell(row=int(row), column=1).number_format = 'YYYY-MM-DD'
        # Format hours column with 1 decimal place
        ws.cell(row=int(row), column=5).number_format = '0.0'
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    
    # Save workbook
    wb.save(output_file)
    
    # Print summary
    total_shifts = len(layer_shifts)
    unique_people = len(person_colors)
    days_covered = (end_date - start_date).days
    
    print(f"✓ PagerDuty-style on-call schedule generated: {output_file}")
    print(f"  - Schedule: {schedule_name}")
    print(f"  - Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')} ({days_covered} days)")
    print(f"  - Total layers: {len(layers_config)}")
    print(f"  - Total team members: {unique_people}")
    print(f"  - Total shifts: {total_shifts}")
    
    # Return data for visual generation
    return layer_shifts, person_colors, start_date, end_date, schedule_name


def generate_visual_schedule(layer_shifts, person_colors, start_date, end_date, schedule_name, output_file):
    """
    Generate a visual representation of the on-call schedule.
    
    Args:
        layer_shifts: List of (date, layer_name, time_window, person, layer_idx)
        person_colors: Dictionary mapping person names to color codes
        start_date: Start date
        end_date: End date
        schedule_name: Schedule name
        output_file: Output image filename
    """
    # Show all days - remove limit
    days_to_show = (end_date - start_date).days
    viz_end_date = end_date
    
    # Filter shifts for visualization period
    viz_shifts = [(d, ln, tw, p, li) for d, ln, tw, p, li in layer_shifts if d < viz_end_date]
    
    if not viz_shifts:
        print("  ! No shifts to visualize")
        return
    
    # Group shifts by date
    shifts_by_date = {}
    for shift_date, layer_name, time_window, person, layer_idx in viz_shifts:
        date_key = shift_date.strftime('%Y-%m-%d')
        if date_key not in shifts_by_date:
            shifts_by_date[date_key] = []
        
        # Parse time window
        start_time, end_time = time_window.split(' - ')
        shifts_by_date[date_key].append({
            'layer': layer_name,
            'start': start_time,
            'end': end_time,
            'person': person,
            'date': shift_date
        })
    
    # Create figure
    dates = sorted(shifts_by_date.keys())
    num_dates = len(dates)
    
    # Calculate figure size based on number of days
    # Adjust width per day for better scaling with many days
    fig_width = max(16, num_dates * 1.2)  # Increased from 0.4 to 1.2 (x3 wider)
    fig_height = 10
    
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    
    # Convert hex colors to RGB
    def hex_to_rgb(hex_color):
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) / 255.0 for i in (0, 2, 4))
    
    # Time to y-coordinate mapping
    def time_to_y(time_str):
        hours, minutes = map(int, time_str.split(':'))
        return hours + minutes / 60.0
    
    # Find global min and max times across all shifts for dynamic Y-axis
    all_times = []
    for date_str in dates:
        for shift in shifts_by_date[date_str]:
            all_times.append(time_to_y(shift['start']))
            all_times.append(time_to_y(shift['end']))
    
    min_time = int(min(all_times))
    max_time = int(max(all_times)) + 1
    
    # Plot shifts
    bar_width = 0.8
    x_pos = 0
    
    date_positions = {}
    for date_str in dates:
        shifts = shifts_by_date[date_str]
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        day_name = date_obj.strftime('%a')
        
        date_positions[date_str] = x_pos
        
        for shift in shifts:
            y_start = time_to_y(shift['start'])
            y_end = time_to_y(shift['end'])
            height = y_end - y_start
            
            person = shift['person']
            color_hex = person_colors.get(person, 'CCCCCC')
            color_rgb = hex_to_rgb(color_hex)
            
            # Draw rectangle for shift
            rect = Rectangle((x_pos - bar_width/2, y_start), bar_width, height,
                           facecolor=color_rgb, edgecolor='black', linewidth=1)
            ax.add_patch(rect)
            
            # Add person name in the middle of the bar (horizontal text)
            text_y = y_start + height / 2
            ax.text(x_pos, text_y, person, ha='center', va='center',
                   fontsize=7, fontweight='bold', rotation=0)
        
        # Add date label above (position above the chart area) - horizontal text
        ax.text(x_pos, min_time - 1.5, f"{date_str}\n{day_name}", ha='center', va='top',
               fontsize=7, fontweight='bold', rotation=0)
        
        x_pos += 1
    
    # Configure axes
    ax.set_xlim(-0.5, num_dates - 0.5)
    ax.set_ylim(min_time - 2, max_time)  # Increased space from -1 to -2 for dates
    ax.invert_yaxis()  # Earlier times at top
    
    # Y-axis: hours (dynamic based on actual shift times)
    ax.set_yticks(range(min_time, max_time))
    ax.set_yticklabels([f"{h:02d}:00" for h in range(min_time, max_time)])
    ax.set_ylabel('Time', fontsize=12, fontweight='bold')
    
    # X-axis: remove ticks (dates shown on plot)
    ax.set_xticks([])
    ax.set_xlabel('Date', fontsize=12, fontweight='bold')
    
    # Grid
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    
    # Title
    title = f"{schedule_name}\n{start_date.strftime('%Y-%m-%d')} to {viz_end_date.strftime('%Y-%m-%d')}"
    ax.set_title(title, fontsize=14, fontweight='bold', pad=20)
    
    # Legend for people
    legend_elements = []
    for person in sorted(person_colors.keys()):
        color_hex = person_colors[person]
        color_rgb = hex_to_rgb(color_hex)
        legend_elements.append(mpatches.Patch(facecolor=color_rgb, edgecolor='black', label=person))
    
    ax.legend(handles=legend_elements, loc='upper right', bbox_to_anchor=(1.12, 1),
             title='Team Members', fontsize=9)
    
    plt.tight_layout()
    plt.savefig(output_file, dpi=150, bbox_inches='tight')
    plt.close()
    
    print(f"✓ Visual schedule generated: {output_file}")
    print(f"  - Showing all {days_to_show} days")


def generate_ics_files(layer_shifts, schedule_name, output_dir="."):
    """
    Generate ICS (iCalendar) files for each person in the schedule.
    
    Args:
        layer_shifts: List of (date, layer_name, time_window, person, layer_idx)
        schedule_name: Schedule name
        output_dir: Directory to save ICS files
    """
    from collections import defaultdict
    
    # Group shifts by person
    shifts_by_person = defaultdict(list)
    for shift_date, layer_name, time_window, person, layer_idx in layer_shifts:
        shifts_by_person[person].append((shift_date, layer_name, time_window))
    
    # Create output directory if it doesn't exist
    ics_dir = os.path.join(output_dir, "ics_files")
    os.makedirs(ics_dir, exist_ok=True)
    
    # Generate ICS file for each person
    for person, shifts in shifts_by_person.items():
        # Sanitize filename and pad numbers with zero (e.g., Utente 1 -> Utente 01)
        import re
        safe_filename = "".join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in person)
        # Replace single digit numbers with zero-padded version
        safe_filename = re.sub(r'\b(\d)\b', r'0\1', safe_filename)
        ics_file = os.path.join(ics_dir, f"{safe_filename}.ics")
        
        with open(ics_file, 'w', encoding='utf-8') as f:
            # Write ICS header
            f.write("BEGIN:VCALENDAR\n")
            f.write("VERSION:2.0\n")
            f.write("PRODID:-//On-Call Scheduler//EN\n")
            f.write(f"X-WR-CALNAME:{person} - On-Call Schedule\n")
            f.write("X-WR-TIMEZONE:UTC\n")
            f.write("CALSCALE:GREGORIAN\n")
            f.write("METHOD:PUBLISH\n")
            
            # Write each shift as an event
            for shift_date, layer_name, time_window in shifts:
                # Parse time window
                time_parts = time_window.split(' - ')
                start_time_str = time_parts[0] if len(time_parts) > 0 else '00:00'
                end_time_str = time_parts[1] if len(time_parts) > 1 else '23:59'
                
                # Parse times
                start_hour, start_min = map(int, start_time_str.split(':'))
                end_hour, end_min = map(int, end_time_str.split(':'))
                
                # Create datetime objects
                start_dt = shift_date.replace(hour=start_hour, minute=start_min, second=0, microsecond=0)
                end_dt = shift_date.replace(hour=end_hour, minute=end_min, second=0, microsecond=0)
                
                # Format for ICS (YYYYMMDDTHHMMSS)
                start_str = start_dt.strftime('%Y%m%dT%H%M%S')
                end_str = end_dt.strftime('%Y%m%dT%H%M%S')
                
                # Generate unique UID
                uid = f"{start_str}-{person.replace(' ', '-')}-oncall@scheduler"
                
                # Current timestamp for DTSTAMP
                now_str = datetime.now().strftime('%Y%m%dT%H%M%SZ')
                
                # Write event
                f.write("BEGIN:VEVENT\n")
                f.write(f"UID:{uid}\n")
                f.write(f"DTSTAMP:{now_str}\n")
                f.write(f"DTSTART:{start_str}\n")
                f.write(f"DTEND:{end_str}\n")
                f.write(f"SUMMARY:On-Call: {layer_name}\n")
                f.write(f"DESCRIPTION:On-call shift for {person}\\nLayer: {layer_name}\\nSchedule: {schedule_name}\n")
                f.write(f"LOCATION:On-Call\n")
                f.write("STATUS:CONFIRMED\n")
                f.write("TRANSP:OPAQUE\n")
                f.write("BEGIN:VALARM\n")
                f.write("TRIGGER:-PT15M\n")
                f.write("ACTION:DISPLAY\n")
                f.write("DESCRIPTION:On-Call shift starts in 15 minutes\n")
                f.write("END:VALARM\n")
                f.write("END:VEVENT\n")
            
            # Write ICS footer
            f.write("END:VCALENDAR\n")
    
    print(f"✓ ICS files generated in: {ics_dir}/")
    print(f"  - Generated {len(shifts_by_person)} calendar files")
    for person in sorted(shifts_by_person.keys()):
        import re
        safe_filename = "".join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in person)
        # Replace single digit numbers with zero-padded version
        safe_filename = re.sub(r'\b(\d)\b', r'0\1', safe_filename)
        print(f"    • {safe_filename}.ics ({len(shifts_by_person[person])} shifts)")



def main():
    parser = argparse.ArgumentParser(
        description='Generate PagerDuty-style on-call shift schedule in xlsx format',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Configuration Files:
  Use YAML configuration files to define complete schedules with multiple layers.
  Each layer represents a single time window (e.g., 08:00-10:30) that rotates daily.
  All layers together compose the complete shift coverage.
  
  Team members are defined within the YAML configuration file, not on command line.
  
  Example configurations included:
    - layers_2_5h_shifts.yaml: 7 layers (4 for Mon-Thu 08:00-18:00, 3 for Fri 08:00-13:00)

Examples:
  %(prog)s --config layers_2_5h_shifts.yaml
  %(prog)s --config layers_2_5h_shifts.yaml --output my_schedule.xlsx
  %(prog)s --config layers_2_5h_shifts.yaml --start-date 2026-02-01 --end-date 2026-05-01
        """
    )
    
    parser.add_argument(
        '--config',
        required=True,
        help='Path to YAML configuration file'
    )
    
    parser.add_argument(
        '--start-date',
        help='Start date: YYYY-MM-DD (e.g. 2026-01-20), relative (e.g. +2w, +3m), or "today"'
    )
    
    parser.add_argument(
        '--end-date',
        help='End date: YYYY-MM-DD, relative from start (e.g. +2w, +3m), or relative from today'
    )
    
    parser.add_argument(
        '--generate-ics',
        action='store_true',
        help='Generate ICS (iCalendar) files for each team member'
    )
    
    args = parser.parse_args()
    
    # Create output directory and filename based on config file name
    config_basename = os.path.splitext(os.path.basename(args.config))[0]
    output_dir = config_basename
    os.makedirs(output_dir, exist_ok=True)
    
    # Auto-generate output filename from config name
    output_file = os.path.join(output_dir, f"{config_basename}.xlsx")
    
    # Parse dates
    start_date = None
    end_date = None
    
    if args.start_date:
        try:
            start_date = parse_date_argument(args.start_date)
            print(f"  Start date: {start_date.strftime('%Y-%m-%d')} ({args.start_date})")
        except ValueError as e:
            print(f"Error: {e}")
            return 1
    
    if args.end_date:
        try:
            # For relative end dates, calculate from start_date if provided, otherwise from today
            reference = start_date if start_date else datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = parse_date_argument(args.end_date, reference)
            print(f"  End date: {end_date.strftime('%Y-%m-%d')} ({args.end_date})")
        except ValueError as e:
            print(f"Error: {e}")
            return 1
    
    # Generate the schedule
    try:
        result = generate_oncall_calendar(args.config, output_file, start_date, end_date)
        
        if result:
            layer_shifts, person_colors, start_date_actual, end_date_actual, schedule_name = result
            
            # Generate visual representation
            visual_output = output_file.replace('.xlsx', '.png')
            generate_visual_schedule(layer_shifts, person_colors, start_date_actual, 
                                   end_date_actual, schedule_name, visual_output)
            
            # Generate ICS files if requested
            if args.generate_ics:
                generate_ics_files(layer_shifts, schedule_name, output_dir)
        
    except FileNotFoundError as e:
        print(f"Error: {e}")
        return 1
    except ValueError as e:
        print(f"Error: {e}")
        return 1
    except Exception as e:
        print(f"Error generating schedule: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == '__main__':
    exit(main())
